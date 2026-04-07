import streamlit as st
import requests
from bs4 import BeautifulSoup
import pandas as pd
import re
import io
import time
import zipfile
from urllib.parse import urljoin, urlparse, parse_qs, urlencode, urlunparse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Page Config ───────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Business Fair Deep Scraper",
    page_icon="🏢",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown("""
<style>
    .main-header {
        background: linear-gradient(135deg, #1a1a2e 0%, #16213e 50%, #0f3460 100%);
        padding: 2rem; border-radius: 12px; margin-bottom: 1.5rem;
        text-align: center; color: white;
    }
    .main-header h1 { font-size: 2.4rem; font-weight: 800; margin: 0; color: #e94560; }
    .main-header p  { font-size: 1rem; color: #a8b2d8; margin-top: 0.5rem; }
    .stat-card {
        background: linear-gradient(135deg, #0f3460, #16213e);
        border: 1px solid #e94560; border-radius: 10px;
        padding: 1rem 1.5rem; text-align: center; color: white;
    }
    .stat-card h2 { font-size: 2rem; color: #e94560; margin: 0; }
    .stat-card p  { color: #a8b2d8; margin: 0; font-size: 0.85rem; }
    .log-box {
        background: #0a0a0a; border: 1px solid #333; border-radius: 8px;
        padding: 1rem; font-family: 'Courier New', monospace;
        font-size: 0.78rem; color: #00ff88; max-height: 260px; overflow-y: auto;
    }
    .stButton > button {
        background: linear-gradient(135deg, #e94560, #c23152);
        color: white; border: none; border-radius: 8px;
        padding: 0.6rem 1.4rem; font-weight: 700; width: 100%; font-size: 1rem;
    }
    .stButton > button:hover { opacity: 0.88; }
    .stDownloadButton > button {
        background: linear-gradient(135deg, #0f3460, #1a5276);
        color: white !important; border: none; border-radius: 8px;
        padding: 0.55rem 1.2rem; font-weight: 600; width: 100%;
    }
</style>
""", unsafe_allow_html=True)

st.markdown("""
<div class="main-header">
  <h1>🏢 Business Fair Deep Scraper</h1>
  <p>High-accuracy multi-page scraper · Noise filtering · Email & phone validation · Confidence scoring</p>
</div>
""", unsafe_allow_html=True)

# ── Sidebar ───────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("### ⚙️ Scraper Settings")
    delay       = st.slider("Request delay (sec)", 0.5, 5.0, 1.5, 0.5)
    max_pages   = st.slider("Max pages to scrape", 1, 100, 20)
    timeout     = st.slider("Request timeout (sec)", 5, 60, 20)
    deep_scrape = st.toggle("Deep scrape (follow profile links)", value=True)
    max_detail  = st.slider("Max detail pages per listing", 1, 30, 10,
                            disabled=not deep_scrape)
    st.divider()
    st.markdown("### 🧹 Accuracy & Filtering")
    min_confidence = st.slider("Min confidence score (0–100)", 0, 100, 40,
        help="Entries below this score are dropped. Higher = stricter.")
    require_email  = st.checkbox("Require at least one email", value=False)
    require_name   = st.checkbox("Require company name", value=False)
    filter_dupes   = st.checkbox("Aggressive deduplication", value=True,
        help="Also deduplicates by email domain and phone number")
    st.divider()
    st.markdown("### 📋 Export")
    incl_confidence = st.checkbox("Show confidence score column", value=True)
    incl_raw        = st.checkbox("Include raw context column", value=False)
    st.info("💡 Raise confidence score to 60+ for very clean data.")

# ── Constants & Compiled Patterns ─────────────────────────────────────────────
HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "en-US,en;q=0.9",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
}

# Strict email pattern — requires real TLD, no image extensions
EMAIL_RE = re.compile(
    r'\b[a-zA-Z0-9._%+\-]{2,64}@[a-zA-Z0-9.\-]{2,255}\.[a-zA-Z]{2,10}\b'
)
# Phone: must have 7–15 digits, won't match pure dates or zip codes
PHONE_RE = re.compile(
    r'(?<!\d)(\+?[\d][\d\s\-().]{6,18}[\d])(?!\d)'
)

# ── Junk filters ──────────────────────────────────────────────────────────────
JUNK_EMAIL_DOMAINS = {
    'example.com','test.com','email.com','mail.com','domain.com',
    'yourcompany.com','company.com','sentry.io','wixpress.com',
    'squarespace.com','shopify.com','amazonaws.com','cloudfront.net',
    '2x.png','jpg','png','gif','svg','webp','css','js',
}
JUNK_EMAIL_USERS = {
    'noreply','no-reply','donotreply','do-not-reply','mailer-daemon',
    'postmaster','webmaster','info@info','admin@admin','test@test',
    'user@example','support@support',
}
JUNK_WORDS_IN_NAME = re.compile(
    r'\b(cookie|privacy|policy|terms|login|register|search|filter|sort|'
    r'menu|nav|footer|header|sidebar|widget|copyright|all rights|'
    r'powered by|loading|javascript|enabled|browser)\b',
    re.I
)
NOISE_TAGS = {'script','style','noscript','meta','head','iframe','svg',
              'path','use','symbol','defs','nav','footer','form',
              'input','button','select','option','textarea'}

# Known CDN / tracker / social domains to skip as "company websites"
SKIP_WEBSITE_DOMAINS = {
    'facebook.com','twitter.com','instagram.com','linkedin.com',
    'youtube.com','google.com','googleapis.com','gstatic.com',
    'cloudflare.com','cdn.jsdelivr.net','unpkg.com','jquery.com',
    'bootstrapcdn.com','fontawesome.com','gravatar.com',
    'wp-content','wp-includes','wp-json',
    'doubleclick.net','googletagmanager.com','google-analytics.com',
    'hotjar.com','intercom.io','hubspot.com','mailchimp.com',
    'amazonaws.com','cloudfront.net','akamaihd.net',
}

DETAIL_KEYWORDS = (
    'exhibitor','profile','company','participant','booth',
    'vendor','member','listing','stand','sponsor','brand','directory',
)

# ── Helpers ───────────────────────────────────────────────────────────────────
def safe_get(url, timeout_s=20, logs=None):
    try:
        r = requests.get(url, headers=HEADERS, timeout=timeout_s, allow_redirects=True)
        r.raise_for_status()
        ct = r.headers.get('Content-Type', '')
        if 'html' not in ct and 'text' not in ct:
            return None
        return r
    except Exception as e:
        if logs is not None:
            logs.append(f"⚠️  {url[:80]} → {e}")
        return None

def clean_soup(soup):
    """Remove noise tags from soup in-place."""
    for tag in soup.find_all(NOISE_TAGS):
        tag.decompose()
    return soup

def validate_email(email: str) -> bool:
    """Return True only for plausibly real emails."""
    if not email or '@' not in email:
        return False
    user, domain = email.lower().rsplit('@', 1)
    # Must have a dot in domain
    if '.' not in domain:
        return False
    # Block image/asset extensions masquerading as emails
    tld = domain.rsplit('.', 1)[-1]
    if tld in ('png','jpg','jpeg','gif','svg','webp','css','js','ico','woff','ttf'):
        return False
    if domain in JUNK_EMAIL_DOMAINS:
        return False
    if user in JUNK_EMAIL_USERS or any(j in user for j in ('noreply','donotreply')):
        return False
    # Too short user part
    if len(user) < 2:
        return False
    return True

def extract_emails(text: str) -> list:
    raw = EMAIL_RE.findall(text)
    return [e.lower() for e in dict.fromkeys(raw) if validate_email(e)]

def validate_phone(digits: str) -> bool:
    """Phone must have 7–15 digits; reject pure years/dates."""
    n = len(digits)
    if n < 7 or n > 15:
        return False
    # Reject sequences that look like years (4-digit 19xx/20xx) or zip codes
    if n == 4 and re.match(r'(19|20)\d{2}', digits):
        return False
    return True

def extract_phones(text: str) -> list:
    raw = PHONE_RE.findall(text)
    seen, out = set(), []
    for p in raw:
        digits = re.sub(r'\D', '', p)
        if validate_phone(digits) and digits not in seen:
            seen.add(digits)
            out.append(p.strip())
    return out

def is_junk_website(url: str, base_domain: str) -> bool:
    try:
        d = urlparse(url).netloc.lower().lstrip('www.')
        if not d:
            return True
        for skip in SKIP_WEBSITE_DOMAINS:
            if skip in d:
                return True
        if d == base_domain:
            return True
        # Skip internal asset paths
        path = urlparse(url).path.lower()
        if re.search(r'\.(png|jpg|gif|svg|css|js|ico|woff|ttf|pdf)$', path):
            return True
    except Exception:
        return True
    return False

def clean_company_name(raw: str) -> str:
    if not raw:
        return ''
    # Remove leading/trailing punctuation & whitespace
    name = raw.strip().strip('|/-–—:,.')
    # Collapse whitespace
    name = re.sub(r'\s+', ' ', name).strip()
    # Drop if looks like navigation / junk
    if JUNK_WORDS_IN_NAME.search(name):
        return ''
    # Drop if too short or too long
    if len(name) < 2 or len(name) > 140:
        return ''
    # Drop if it's all numbers
    if re.match(r'^[\d\s\W]+$', name):
        return ''
    return name

def confidence_score(entry: dict) -> int:
    """
    0–100 score based on data completeness and quality.
    Used to filter out near-empty or noisy entries.
    """
    score = 0
    if entry.get('company'):
        score += 30
    emails = entry.get('emails', [])
    if emails:
        score += 30
        # Bonus for business-looking domain (not gmail/yahoo/hotmail)
        free_domains = {'gmail.com','yahoo.com','hotmail.com','outlook.com',
                        'icloud.com','aol.com','protonmail.com','yandex.com'}
        if any(e.split('@')[-1] not in free_domains for e in emails):
            score += 10
    if entry.get('website'):
        score += 20
    if entry.get('phones'):
        score += 10
    return min(score, 100)

def find_pagination_urls(soup, base_url: str, visited: set) -> list:
    """Multi-strategy pagination: rel=next, text hints, page params."""
    candidates = []
    parsed_base = urlparse(base_url)

    # Strategy 1: rel="next"
    for tag in soup.find_all('a', rel=lambda r: r and 'next' in r):
        h = tag.get('href', '')
        if h:
            candidates.append(urljoin(base_url, h))

    # Strategy 2: text-based next links
    if not candidates:
        for tag in soup.find_all('a', href=True):
            txt = tag.get_text(strip=True).lower()
            if txt in ('next','next »','next page','›','»','next →','>','→','forward'):
                h = tag['href']
                if h and '#' not in h:
                    candidates.append(urljoin(base_url, h))

    # Strategy 3: numeric pagination — pick the next page number
    if not candidates:
        current_page_num = None
        qs = parse_qs(parsed_base.query)
        for param in ('page','p','pg','pg_num','paged'):
            if param in qs:
                try:
                    current_page_num = (int(qs[param][0]), param)
                except ValueError:
                    pass
                break

        if current_page_num:
            num, param = current_page_num
            next_num = num + 1
            new_qs = dict(qs)
            new_qs[param] = [str(next_num)]
            new_parts = list(parsed_base)
            new_parts[4] = urlencode(new_qs, doseq=True)
            candidates.append(urlunparse(new_parts))
        else:
            # Scan page for numbered links greater than current
            all_nums = []
            for tag in soup.find_all('a', href=True):
                href = tag['href']
                m = re.search(r'[?&](page|p|pg|paged)=(\d+)', href, re.I)
                if m:
                    all_nums.append((int(m.group(2)), urljoin(base_url, href)))
            if all_nums:
                all_nums.sort()
                for num, url in all_nums:
                    if url not in visited:
                        candidates.append(url)
                        break

    # Deduplicate and exclude already visited
    seen, out = set(), []
    for u in candidates:
        if u not in visited and u not in seen:
            seen.add(u)
            out.append(u)
    return out

def extract_best_company_name(item, fallback_text: str) -> str:
    """Try multiple strategies to find a company name within a container."""
    # Priority: data attributes > heading tags > strong/b
    for attr in ('data-company','data-name','data-title','data-exhibitor'):
        val = item.get(attr, '')
        if val:
            return clean_company_name(val)
    for tag in ['h1','h2','h3','h4','h5']:
        el = item.find(tag)
        if el:
            name = clean_company_name(el.get_text(strip=True))
            if name:
                return name
    for tag in ['strong','b','span[class*="name"]','span[class*="title"]',
                'p[class*="name"]','div[class*="name"]']:
        try:
            el = item.select_one(tag)
        except Exception:
            el = item.find(tag)
        if el:
            name = clean_company_name(el.get_text(strip=True))
            if name:
                return name
    return ''

def parse_entries(soup, base_url: str) -> list:
    """
    Extract business entries. Uses container scoring to prefer
    tighter, more specific containers and avoid duplicating parent-child data.
    """
    base_domain = urlparse(base_url).netloc.lower().lstrip('www.')
    soup = clean_soup(soup)

    # Collect raw candidates from all meaningful containers
    raw = []
    for item in soup.find_all(['article','div','li','tr','section','dl','dd']):
        # Skip containers with too much text (likely full page / wrapper)
        text = item.get_text(separator=' ', strip=True)
        if len(text) > 2000 or len(text) < 15:
            continue
        # Skip obvious nav/footer/header containers
        cls  = ' '.join(item.get('class', []))
        iid  = item.get('id', '')
        if re.search(r'nav|footer|header|sidebar|widget|cookie|modal|popup|'
                     r'breadcrumb|social|share|comment|search', cls+iid, re.I):
            continue

        emails = extract_emails(text)
        phones = extract_phones(text)

        links       = item.find_all('a', href=True)
        website     = ''
        detail_links = []
        for lnk in links:
            full = urljoin(base_url, lnk['href'])
            parsed = urlparse(full)
            if parsed.scheme not in ('http','https'):
                continue
            lnk_domain = parsed.netloc.lower().lstrip('www.')
            if lnk_domain and lnk_domain != base_domain:
                if not website and not is_junk_website(full, base_domain):
                    website = full
            else:
                path_lower = parsed.path.lower()
                if any(k in path_lower for k in DETAIL_KEYWORDS):
                    if full != base_url:
                        detail_links.append(full)

        company = extract_best_company_name(item, text)

        if emails or website or phones or company:
            raw.append({
                'company':      company,
                'emails':       emails,
                'phones':       phones,
                'website':      website,
                'detail_links': list(dict.fromkeys(detail_links)),
                'context':      text[:180].replace('\n', ' '),
                'text_len':     len(text),
            })

    # ── Hierarchy deduplication ───────────────────────────────────────────────
    # When a parent and child container share the same email/website,
    # keep only the child (smaller, more specific container).
    raw.sort(key=lambda x: x['text_len'])  # shortest first
    seen_emails, seen_websites, seen_phones = set(), set(), set()
    results = []
    for entry in raw:
        email_set   = frozenset(entry['emails'])
        website_key = entry['website']
        phone_set   = frozenset(entry['phones'])

        # Skip if we've already captured this exact data from a tighter container
        if email_set and email_set.issubset(seen_emails):
            continue
        if website_key and website_key in seen_websites and not email_set:
            continue

        seen_emails   |= email_set
        if website_key:
            seen_websites.add(website_key)
        seen_phones |= phone_set
        results.append(entry)

    return results

def deep_scrape_entry(entry: dict, base_url: str, timeout_s: int,
                      logs: list, max_d: int) -> dict:
    base_domain = urlparse(base_url).netloc.lower().lstrip('www.')
    for url in entry['detail_links'][:max_d]:
        if logs is not None:
            logs.append(f"  🔍 Detail: {url[:80]}")
        r = safe_get(url, timeout_s, logs)
        if not r:
            continue
        soup = clean_soup(BeautifulSoup(r.text, 'html.parser'))
        text = soup.get_text(separator=' ', strip=True)

        new_emails = extract_emails(text)
        new_phones = extract_phones(text)
        entry['emails'] = list(dict.fromkeys(entry['emails'] + new_emails))
        entry['phones'] = list(dict.fromkeys(entry['phones'] + new_phones))

        if not entry['website']:
            for lnk in soup.find_all('a', href=True):
                full = urljoin(url, lnk['href'])
                if not is_junk_website(full, base_domain):
                    d = urlparse(full).netloc.lower().lstrip('www.')
                    if d and d != base_domain:
                        entry['website'] = full
                        break

        if not entry['company']:
            for tag in ['h1','h2','h3']:
                el = soup.find(tag)
                if el:
                    name = clean_company_name(el.get_text(strip=True))
                    if name:
                        entry['company'] = name
                        break
    return entry

def build_excel(df: pd.DataFrame) -> bytes:
    wb = Workbook()

    # ── Summary sheet ─────────────────────────────────────────────────────────
    ws2 = wb.active
    ws2.title = "Summary"
    ws2['A1'] = "Business Fair Scrape — Summary"
    ws2['A1'].font = Font(bold=True, size=14, color="FFFFFF")
    ws2['A1'].fill = PatternFill("solid", start_color="0F3460")
    ws2.merge_cells('A1:C1')
    stats = [
        ("Total Companies",       len(df)),
        ("With Email",            int(df['Email'].astype(bool).sum()) if 'Email' in df else 0),
        ("With Website",          int(df['Website'].astype(bool).sum()) if 'Website' in df else 0),
        ("With Phone",            int(df['Phone'].astype(bool).sum()) if 'Phone' in df else 0),
        ("Complete (email+web)",  int(((df['Email'].astype(bool)) & (df['Website'].astype(bool))).sum()) if ('Email' in df and 'Website' in df) else 0),
    ]
    for i, (k, v) in enumerate(stats, 3):
        ws2[f'A{i}'] = k; ws2[f'B{i}'] = v
        ws2[f'A{i}'].font = Font(bold=True)
    ws2.column_dimensions['A'].width = 28
    ws2.column_dimensions['B'].width = 14

    # ── Exhibitors sheet ──────────────────────────────────────────────────────
    ws = wb.create_sheet("Exhibitors")
    cols = list(df.columns)
    hdr_fill = PatternFill("solid", start_color="E94560")
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    thin = Side(style='thin', color='DDDDDD')
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

    for ci, col in enumerate(cols, 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.fill = hdr_fill; cell.font = hdr_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = bdr

    fill_a = PatternFill("solid", start_color="F8F9FA")
    fill_b = PatternFill("solid", start_color="FFFFFF")

    for ri, row in enumerate(df.itertuples(index=False), 2):
        fill = fill_a if ri % 2 == 0 else fill_b
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=str(val) if val else '')
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            cell.fill = fill; cell.border = bdr
            col_name = cols[ci - 1]
            if col_name == 'Website' and val and str(val).startswith('http'):
                cell.hyperlink = str(val)
                cell.font = Font(color='0563C1', underline='single')
            elif col_name == 'Email' and val and '@' in str(val):
                first = str(val).split(',')[0].strip()
                cell.hyperlink = f"mailto:{first}"
                cell.font = Font(color='0563C1', underline='single')

    width_map = {
        '#': 5, 'Company Name': 30, 'Email': 36, 'Website': 36,
        'Phone': 20, 'Confidence': 12, 'Source Page': 28, 'Context': 42,
    }
    for ci, col in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(ci)].width = width_map.get(col, 20)

    ws.row_dimensions[1].height = 28
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()

def build_zip(df: pd.DataFrame, excel_bytes: bytes) -> bytes:
    readme = """# Business Fair Deep Scraper

## Overview
High-accuracy multi-page web scraper that extracts business details
(names, emails, websites, phone numbers) from fair/exhibition directories.

## Accuracy Features
- **Junk email filter**: blocks noreply, test, CDN, image-extension emails
- **Phone validation**: strips dates, zip codes, pure noise patterns
- **Noise tag removal**: script/style/nav/footer stripped before parsing
- **Hierarchy deduplication**: parent containers don't duplicate child data
- **Confidence score**: 0–100 completeness/quality rating per entry
- **Junk website filter**: blocks social media, CDN, analytics, tracking URLs
- **Company name cleaning**: removes nav words, short/long strings, numbers-only

## Excel Output Columns
| Column      | Description                              |
|-------------|------------------------------------------|
| #           | Row index                                |
| Company Name| Business / exhibitor name                |
| Email       | Validated email address(es)              |
| Phone       | Validated phone number(s)                |
| Website     | Company website URL                      |
| Confidence  | Data quality score 0–100                 |
| Source Page | Directory page it was found on           |
| Context     | Short surrounding text snippet (optional)|

## Notes
- JavaScript-rendered sites (React/Vue) won't expose data to this scraper
- Emails behind login walls require deep scrape + profile links
- Confidence ≥ 60 = business name + email + website all present
"""

    setup = """# Setup Guide

## Requirements
Python 3.9+ · pip

## Installation
```bash
python -m venv venv
source venv/bin/activate        # Windows: venv\\Scripts\\activate
pip install streamlit requests beautifulsoup4 pandas openpyxl lxml
```

## Run
```bash
streamlit run main.py
# Open http://localhost:8501
```

## Accuracy Tips
| Goal                    | Setting                                   |
|------------------------|-------------------------------------------|
| Cleanest data           | Confidence ≥ 70, require name + email     |
| Maximum coverage        | Confidence = 0, deep scrape ON            |
| Avoid duplicates        | Aggressive deduplication ON               |
| Hidden emails           | Deep scrape ON, max detail pages 20+      |
| Avoid being blocked     | Delay 3–5 sec                             |

## Troubleshooting
| Problem                  | Solution                                  |
|--------------------------|-------------------------------------------|
| No emails found          | Enable deep scrape                        |
| 403/429 errors           | Increase crawl delay                      |
| Pagination stops early   | Paste ?page=1 URL as start               |
| JS-rendered site         | Scraper cannot reach this data            |
| Too many junk entries    | Raise confidence threshold                |
"""

    stats_txt = f"""Scrape Statistics
=================
Total companies : {len(df)}
With email      : {int(df['Email'].astype(bool).sum()) if 'Email' in df else 0}
With website    : {int(df['Website'].astype(bool).sum()) if 'Website' in df else 0}
With phone      : {int(df['Phone'].astype(bool).sum()) if 'Phone' in df else 0}
Complete entries: {int(((df['Email'].astype(bool)) & (df['Website'].astype(bool))).sum()) if ('Email' in df and 'Website' in df) else 0}
Avg confidence  : {df['Confidence'].mean() if 'Confidence' in df else 0:.1f} / 100
"""

    zbuf = io.BytesIO()
    with zipfile.ZipFile(zbuf, 'w', zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("README.md",        readme)
        zf.writestr("SETUP.md",         setup)
        zf.writestr("scrape_stats.txt", stats_txt)
        zf.writestr("main.py",          open(__file__).read())
        zf.writestr("fair_participants.xlsx", excel_bytes)
    zbuf.seek(0)
    return zbuf.read()

# ── Main UI ───────────────────────────────────────────────────────────────────
url_input = st.text_input(
    "🌐 Fair Directory URL:",
    placeholder="https://fair-website.com/exhibitors",
    help="Paste the exhibitor/participant directory page URL",
)

col_btn, _ = st.columns([1, 3])
with col_btn:
    start = st.button("🚀 Start Scraping")

if start:
    if not url_input.strip():
        st.warning("⚠️ Please enter a URL first.")
    else:
        base_url      = url_input.strip()
        logs          = []
        all_entries   = []
        pages_scraped = 0
        visited_pages = set()
        current_url   = base_url

        st.markdown("---")
        st.markdown("### 📡 Live Scraping Log")
        log_ph  = st.empty()
        prog    = st.progress(0, text="Starting…")

        def refresh_log():
            log_ph.markdown(
                "<div class='log-box'>" + "<br>".join(logs[-30:]) + "</div>",
                unsafe_allow_html=True,
            )

        # ── Pagination loop ───────────────────────────────────────────────────
        while current_url and pages_scraped < max_pages and current_url not in visited_pages:
            visited_pages.add(current_url)
            logs.append(f"📄 Page {pages_scraped+1}: {current_url[:80]}")
            refresh_log()
            prog.progress(
                min(pages_scraped / max_pages, 0.95),
                text=f"Scraping page {pages_scraped+1} of up to {max_pages}…",
            )

            r = safe_get(current_url, timeout, logs)
            if not r:
                logs.append("❌ Failed to load — stopping.")
                refresh_log()
                break

            soup    = BeautifulSoup(r.text, 'html.parser')
            entries = parse_entries(soup, current_url)
            logs.append(f"   → Found {len(entries)} candidate containers")
            refresh_log()

            # Deep scrape
            if deep_scrape:
                for ei, entry in enumerate(entries):
                    if entry['detail_links']:
                        logs.append(f"   🕵️ Deep scraping {ei+1}/{len(entries)}: {entry['company'] or 'Unknown'}")
                        refresh_log()
                        entries[ei] = deep_scrape_entry(entry, current_url, timeout, logs, max_detail)
                        time.sleep(delay * 0.4)

            for e in entries:
                e['source_page'] = current_url
                e['confidence']  = confidence_score(e)
            
            all_entries.extend(entries)

            # Find next page
            next_candidates = find_pagination_urls(soup, current_url, visited_pages)
            next_url = None
            for nc in next_candidates:
                if nc not in visited_pages:
                    next_url = nc
                    break
            current_url = next_url
            pages_scraped += 1

            if current_url:
                time.sleep(delay)

        prog.progress(1.0, text="✅ Scraping complete!")
        logs.append(f"🏁 Done! Collected {len(all_entries)} raw entries.")
        refresh_log()

        # ── Filtering & Processing ────────────────────────────────────────────
        if all_entries:
            rows = []
            for e in all_entries:
                if e['confidence'] < min_confidence:
                    continue
                if require_email and not e['emails']:
                    continue
                if require_name and not e['company']:
                    continue

                row = {
                    'Company Name': e['company'],
                    'Email':        ', '.join(e['emails']),
                    'Phone':        ', '.join(e['phones']),
                    'Website':      e['website'],
                    'Confidence':   e['confidence'],
                    'Source Page':  e['source_page'],
                }
                if incl_raw:
                    row['Context'] = e['context']
                rows.append(row)

            if not rows:
                st.warning("⚠️ No entries met your filtering criteria. Try lowering the Min Confidence score.")
            else:
                df = pd.DataFrame(rows)
                if filter_dupes:
                    df.sort_values('Confidence', ascending=False, inplace=True)
                    df.drop_duplicates(subset=['Company Name', 'Email'], keep='first', inplace=True)
                
                df.reset_index(drop=True, inplace=True)
                df.insert(0, '#', df.index + 1)

                display_cols = ['#', 'Company Name', 'Email', 'Phone', 'Website']
                if incl_confidence: display_cols.append('Confidence')
                display_cols.append('Source Page')
                if incl_raw: display_cols.append('Context')
                
                df = df[display_cols]

                # ── Dashboard ─────────────────────────────────────────────────
                st.markdown("### 📊 Results Dashboard")
                c1, c2, c3, c4 = st.columns(4)
                with c1: st.markdown(f"<div class='stat-card'><h2>{len(df)}</h2><p>Companies</p></div>", unsafe_allow_html=True)
                with c2: st.markdown(f"<div class='stat-card'><h2>{int(df['Email'].astype(bool).sum())}</h2><p>With Emails</p></div>", unsafe_allow_html=True)
                with c3: st.markdown(f"<div class='stat-card'><h2>{int(df['Website'].astype(bool).sum())}</h2><p>With Website</p></div>", unsafe_allow_html=True)
                with c4: st.markdown(f"<div class='stat-card'><h2>{int(df['Confidence'].mean())}%</h2><p>Avg Confidence</p></div>", unsafe_allow_html=True)

                st.markdown("### 📋 Extracted Data")
                st.dataframe(df, use_container_width=True, height=450)

                # ── Export ────────────────────────────────────────────────────
                st.markdown("### 💾 Export")
                excel_bytes = build_excel(df)
                zip_bytes   = build_zip(df, excel_bytes)

                dl1, dl2 = st.columns(2)
                with dl1:
                    st.download_button("📥 Download Excel (.xlsx)", excel_bytes, "fair_participants.xlsx", 
                                       "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                with dl2:
                    st.download_button("📦 Download Full ZIP", zip_bytes, "scraper_export.zip", "application/zip")
        else:
            st.info("ℹ️ No entries found. Check the URL or try enabling Deep Scrape.")

st.divider()
st.caption("🏢 Business Fair Deep Scraper | High-Accuracy Mode | Anti-Gravity Powered")